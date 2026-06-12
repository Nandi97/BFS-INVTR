#!/usr/bin/env python3
"""
Beauty First / Beauty Logix — Warehouse Excel → BFS Inventory App

Reads every product sheet from the warehouse Excel file and imports:
  1. Products + brands + suppliers (via /api/import/products)
  2. Opening stock quantities from the QB qty column (via /api/import/stock)

Usage:
  python3 scripts/import_warehouse.py

Requires the Next.js dev server to be running on http://localhost:3000
"""

import re
import sys
import openpyxl
import requests
from pathlib import Path

# ─── Config ──────────────────────────────────────────────────────────────────

EXCEL_PATH = Path("/Users/alvinmaina/Library/CloudStorage/GoogleDrive-order@beautylogix.ca/My Drive/inventory/Wholesale Product Vendors.xlsx WAREHOUSE VERSION.xlsx")
API_BASE   = "http://localhost:3000"

# Sheets that are aggregates / not raw product lists
SKIP_SHEETS = {"Reorder Schedule"}

# Sheets that belong to locations other than the main warehouse.
# Keys must match the sheet name exactly.
LOCATION_MAP = {
    "BF Inverness":  "BF Inverness",
    "BLX Inverness": "BLX Inverness",
}
DEFAULT_LOCATION = "BF Warehouse"

# Brand overrides: if a sheet has no BRAND column, use this name
SHEET_BRAND_DEFAULTS = {
    "Eyelash Extension":          "Eyelash Extension",
    "Eyebrow Queen Pro":          "Eyebrow Queen Pro",
    "Fernandas Facial Supplies":  "Fernandas Facial Supplies",
    "BLX EquipmentFunriture":     "BLX Equipment",
    "BFS EquipmentFurniture":     "BFS Equipment",
    "BF Professional":            "Beauty First",
    "BF Retail":                  "Beauty First",
    "BF Inverness":               "Inverness",
    "BLX Inverness":              "Inverness",
}


# ─── Helpers ─────────────────────────────────────────────────────────────────

def norm(val):
    """Lowercase, collapse whitespace, strip."""
    if val is None:
        return ""
    return re.sub(r"\s+", " ", str(val).lower().replace("\n", " ")).strip()


def find_header_row_index(ws):
    """Return 0-based index of the column-header row (contains PRODUCT NAME or BARCODE)."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        normed = [norm(c) for c in row]
        if any("product name" in c or c == "barcode" for c in normed):
            return i
    return None


def map_columns(header_row):
    """Map canonical key → column index from the header row."""
    cols = {}
    for i, cell in enumerate(header_row):
        n = norm(cell)
        if not n:
            continue
        if "product name" in n:
            cols.setdefault("name", i)
        if n in ("brand",) or n.startswith("brand "):
            cols.setdefault("brand", i)
        if "barcode" in n:
            cols.setdefault("barcode", i)
        if n == "size":
            cols.setdefault("size", i)
        if "professional" in n or n == "p/r" or n.startswith("p(professional"):
            cols.setdefault("type", i)
        if "supplier" in n and "sku" not in n and "health" not in n:
            cols.setdefault("supplier", i)
        # QB qty: any column containing "qty" and ("quickbooks" or "qbo" or "qb")
        if "qty" in n and ("quickbooks" in n or "qbo" in n or "qb" in n):
            cols.setdefault("qty", i)
    return cols


def clean_barcode(raw):
    """Strip float .0 suffix from numeric barcodes."""
    if raw is None:
        return None
    s = str(raw).strip()
    if re.match(r"^\d+\.0$", s):
        s = s[:-2]
    return s if s else None


def resolve_type(raw):
    if not raw:
        return "BOTH"
    v = str(raw).strip().upper().replace(" ", "")
    if v in ("P", "PROFESSIONAL"):
        return "PROFESSIONAL"
    if v in ("R", "RETAIL"):
        return "RETAIL"
    return "BOTH"


def clean_qty(raw):
    """Return non-negative float, or None if unparseable."""
    if raw is None:
        return None
    try:
        v = float(raw)
        return round(max(0.0, v), 2)  # clamp negatives → 0
    except (ValueError, TypeError):
        return None


# ─── Sheet parser ─────────────────────────────────────────────────────────────

def parse_sheet(sheet_name, ws):
    """
    Returns (product_rows, stock_rows) for a single sheet.
    product_rows: list of dicts for /api/import/products
    stock_rows:   list of dicts for /api/import/stock  (only where qty is known)
    """
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    header_idx = find_header_row_index(ws)
    if header_idx is None:
        print(f"  [SKIP] {sheet_name}: could not find header row")
        return [], []

    header_row = all_rows[header_idx]
    cols       = map_columns(header_row)

    # Troiareuke has no PRODUCT NAME — use brand + size as name
    troiareuke_mode = "name" not in cols and "size" in cols

    default_brand = SHEET_BRAND_DEFAULTS.get(sheet_name)
    location      = LOCATION_MAP.get(sheet_name, DEFAULT_LOCATION)

    product_rows = []
    stock_rows   = []

    last_brand = default_brand  # fill-down tracker

    for row in all_rows[header_idx + 1:]:
        # Extract raw values
        def get(key, default=None):
            idx = cols.get(key)
            return row[idx] if idx is not None and idx < len(row) else default

        raw_brand   = get("brand")
        raw_name    = get("name")
        raw_barcode = get("barcode")
        raw_size    = get("size")
        raw_type    = get("type")
        raw_supplier= get("supplier")
        raw_qty     = get("qty")

        # Brand fill-down
        if raw_brand and str(raw_brand).strip():
            last_brand = str(raw_brand).strip()

        # Build product name
        if troiareuke_mode:
            size_str = str(raw_size).strip().replace("\n", " ") if raw_size else ""
            name = f"{last_brand} {size_str}".strip() if size_str else None
        else:
            name = str(raw_name).strip() if raw_name else None

        if not name or name.lower() in ("none", "nan"):
            continue

        # Skip obvious non-product rows
        if any(skip in name.lower() for skip in ["s.no.", "product name", "brand", "total"]):
            continue

        barcode  = clean_barcode(raw_barcode)
        size_val = str(raw_size).strip().replace("\n", " ") if raw_size and not troiareuke_mode else None
        p_type   = resolve_type(raw_type)
        supplier = str(raw_supplier).strip() if raw_supplier and str(raw_supplier).strip() else None
        brand    = last_brand or default_brand

        product_rows.append({
            "name":     name,
            "brand":    brand,
            "barcode":  barcode,
            "size":     size_val,
            "type":     p_type,
            "supplier": supplier,
        })

        qty = clean_qty(raw_qty)
        if qty is not None:
            # Use barcode as identifier if available, else name
            identifier = barcode if barcode else name
            stock_rows.append({
                "identifier": identifier,
                "location":   location,
                "quantity":   qty,
            })

    return product_rows, stock_rows


# ─── API calls ────────────────────────────────────────────────────────────────

def post(endpoint, payload, label):
    try:
        r = requests.post(f"{API_BASE}{endpoint}", json=payload, timeout=120)
        r.raise_for_status()
        return r.json()
    except requests.exceptions.ConnectionError:
        print(f"\n  ERROR: Cannot connect to {API_BASE}. Is the dev server running?")
        sys.exit(1)
    except Exception as e:
        print(f"\n  ERROR {label}: {e}")
        return None


def ensure_location(name, code):
    """Create location if it doesn't exist (ignore 409 conflict)."""
    try:
        r = requests.post(
            f"{API_BASE}/api/locations",
            json={"name": name, "code": code, "type": "WAREHOUSE"},
            timeout=15,
        )
        if r.status_code in (200, 201):
            print(f"  Created location: {name} ({code})")
        elif r.status_code == 409:
            pass  # already exists
        else:
            print(f"  Warning: could not create location {name}: {r.status_code}")
    except Exception as e:
        print(f"  Warning: location creation failed: {e}")


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at:\n  {EXCEL_PATH}")
        sys.exit(1)

    print(f"Reading: {EXCEL_PATH.name}")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)

    # Ensure non-default locations exist
    ensure_location("BF Inverness",  "BFI")
    ensure_location("BLX Inverness", "BLXI")

    all_products = []
    all_stock    = []

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            print(f"  [SKIP] {sheet_name}")
            continue

        ws = wb[sheet_name]
        p_rows, s_rows = parse_sheet(sheet_name, ws)

        print(f"  {sheet_name}: {len(p_rows)} products, {len(s_rows)} with qty")
        all_products.extend(p_rows)
        all_stock.extend(s_rows)

    wb.close()

    print(f"\nTotal: {len(all_products)} product rows, {len(all_stock)} stock rows")

    # ── Step 1: Import products ──────────────────────────────────────────────
    print("\n── Importing products…")

    BATCH = 100
    total_created = total_updated = total_skipped = 0
    all_errors = []

    for i in range(0, len(all_products), BATCH):
        batch  = all_products[i : i + BATCH]
        result = post("/api/import/products", {"rows": batch}, f"products batch {i//BATCH+1}")
        if result:
            total_created  += result.get("created",  0)
            total_updated  += result.get("updated",  0)
            total_skipped  += result.get("skipped",  0)
            all_errors.extend(result.get("errors",  []))
        pct = min(100, round((i + len(batch)) / len(all_products) * 100))
        print(f"  {pct}% — batch {i//BATCH+1} done", end="\r")

    print(f"\n  Products: {total_created} created, {total_updated} updated, {total_skipped} skipped")
    if all_errors:
        print(f"  Product errors ({len(all_errors)}):")
        for e in all_errors[:10]:
            print(f"    {e}")
        if len(all_errors) > 10:
            print(f"    … and {len(all_errors)-10} more")

    # ── Step 2: Import stock ─────────────────────────────────────────────────
    print("\n── Importing stock quantities…")

    total_upserted = total_s_skipped = 0
    stock_errors = []

    for i in range(0, len(all_stock), BATCH):
        batch  = all_stock[i : i + BATCH]
        result = post("/api/import/stock", {"rows": batch}, f"stock batch {i//BATCH+1}")
        if result:
            total_upserted   += result.get("upserted", 0)
            total_s_skipped  += result.get("skipped",  0)
            stock_errors.extend(result.get("errors",  []))
        pct = min(100, round((i + len(batch)) / len(all_stock) * 100))
        print(f"  {pct}% — batch {i//BATCH+1} done", end="\r")

    print(f"\n  Stock: {total_upserted} records set, {total_s_skipped} skipped")
    if stock_errors:
        print(f"  Stock errors ({len(stock_errors)}):")
        for e in stock_errors[:10]:
            print(f"    {e}")
        if len(stock_errors) > 10:
            print(f"    … and {len(stock_errors)-10} more")

    print("\n✓ Import complete. Open http://localhost:3000/stock to see inventory.")


if __name__ == "__main__":
    main()
